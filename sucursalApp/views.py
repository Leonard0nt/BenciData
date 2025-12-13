from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
import calendar
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from typing import Any, Dict, List
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import DecimalField, F, Prefetch, QuerySet, Sum, Value
from django.db.models.functions import Coalesce
from django.http import Http404, HttpResponseRedirect, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.template.defaultfilters import slugify

from django.views import View

from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView
from django.views.generic.edit import FormMixin
from django.views.generic.detail import SingleObjectMixin
from django.views.generic.edit import FormView
from core.mixins import RoleRequiredMixin
from homeApp.models import Company
from UsuarioApp.models import Profile
from .forms import (
    BranchProductForm,
    FuelInventoryForm,
    BranchStaffForm,
    BranchUserLinkForm,
    FuelPriceForm,
    IslandForm,
    MachineForm,
    NozzleForm,
    ServiceSessionCreditSaleForm,
    ServiceSessionFuelLoadForm,
    ServiceSessionFirefighterPaymentForm,
    ServiceSessionProductLoadForm,
    ServiceSessionProductSaleForm,
    ServiceSessionProductSaleItemFormSet,
    ServiceSessionMachineInventoryClosingFormSet,
    ServiceSessionTransbankVoucherForm,
    ServiceSessionWithdrawalForm,
    ShiftForm,
    ServiceSessionForm,
    SucursalForm,
)
from iotApp.models import DispenseEvent
from .models import (
    BranchProduct,
    FuelInventory,
    FuelPrice,
    Island,
    Machine,
    MachineFuelInventoryNumeral,
    Nozzle,
    ServiceSessionCreditSale,
    ServiceSessionFuelLoad,
    ServiceSessionFirefighterPayment,
    ServiceSessionProductLoad,
    ServiceSessionProductSale,
    ServiceSessionProductSaleItem,
    ServiceSessionTransbankVoucher,
    ServiceSessionWithdrawal,
    Shift,
    ServiceSession,
    Sucursal,
    SucursalStaff,
)


def redirect_to_modal(branch_id: int, modal_name: str) -> HttpResponseRedirect:
    """Build a redirect response to reopen a specific modal on the branch page."""

    base_url = reverse("sucursal_update", args=[branch_id])
    query = urlencode({"modal": modal_name})
    return HttpResponseRedirect(f"{base_url}?{query}")


def get_admin_branch_ids(profile) -> List[int]:
    """Return the branch IDs managed by the given administrator profile."""

    if not profile or not getattr(profile, "is_admin", None) or not profile.is_admin():
        return []

    branch_ids = set(
        SucursalStaff.objects.filter(
            profile=profile, role="ADMINISTRATOR"
        ).values_list("sucursal_id", flat=True)
    )

    current_branch_id = getattr(profile, "current_branch_id", None)
    if current_branch_id:
        branch_ids.add(current_branch_id)

    branch_ids.discard(None)
    return list(branch_ids)


class OwnerCompanyMixin(LoginRequiredMixin, RoleRequiredMixin):
    allowed_roles = ["OWNER"]

    def get_company(self) -> Company | None:
        profile = getattr(self.request.user, "profile", None)
        if not profile:
            return None
        try:
            return profile.company
        except Company.DoesNotExist:
            return None
    def get_managed_branch_ids(self) -> List[int]:
        if hasattr(self, "_managed_branch_ids"):
            return self._managed_branch_ids  # type: ignore[attr-defined]

        company = self.get_company()
        if company is not None:
            branch_ids = list(company.branches.values_list("pk", flat=True))
        else:
            profile = getattr(self.request.user, "profile", None)
            # Collect branch ids from SucursalStaff memberships (any role)
            branch_ids_set: set[int] = set()
            if profile:
                staff_branch_ids = list(
                    SucursalStaff.objects.filter(profile=profile).values_list(
                        "sucursal_id", flat=True
                    )
                )
                for bid in staff_branch_ids:
                    if bid:
                        branch_ids_set.add(bid)

                # Also include the profile.current_branch if set
                current_branch_id = getattr(profile, "current_branch_id", None)
                if current_branch_id:
                    branch_ids_set.add(current_branch_id)

                # If the profile is an admin, include admin-managed branches as well
                if getattr(profile, "is_admin", None) and profile.is_admin():
                    for bid in get_admin_branch_ids(profile):
                        if bid:
                            branch_ids_set.add(bid)

            branch_ids = list(branch_ids_set)

        # Ensure unique values and ignore None entries
        branch_ids = [b for i, b in enumerate(branch_ids) if b is not None and b not in branch_ids[:i]]
        self._managed_branch_ids = branch_ids  # type: ignore[attr-defined]
        return branch_ids

    def get_managed_branches_queryset(self) -> QuerySet[Sucursal]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return Sucursal.objects.none()
        return Sucursal.objects.filter(pk__in=branch_ids)
class SucursalListView(OwnerCompanyMixin, FormMixin, ListView):
    model = Sucursal
    template_name = "pages/sucursales/sucursal_list.html"
    context_object_name = "sucursales"
    form_class = SucursalForm
    success_url = reverse_lazy("sucursal_list")
    # Allow owners/admins to edit; also allow other staff to view (read-only)
    allowed_roles = [
        "OWNER",
        "ADMINISTRATOR",
        "ACCOUNTANT",
        "ATTENDANT",
        "HEAD_ATTENDANT",
    ]

    def dispatch(self, request, *args, **kwargs):
        profile = getattr(request.user, "profile", None)

        # Si el usuario **NO** es dueño, mantenemos el redirect a una sucursal
        # excepto cuando es administrador y tiene más de una sucursal asignada
        if profile and not profile.has_role("OWNER"):
            branch_ids = self.get_managed_branch_ids()
            if branch_ids:
                is_admin = profile.has_role("ADMINISTRATOR")
                if is_admin and len(branch_ids) > 1:
                    # Mostrar la lista para que el administrador elija cuál editar
                    return super().dispatch(request, *args, **kwargs)
                # Redirigir a la sucursal asignada
                current_branch_id = getattr(profile, "current_branch_id", None)
                if current_branch_id in branch_ids:
                    branch_id = current_branch_id
                else:
                    branch_id = branch_ids[0]
                return redirect("sucursal_update", pk=branch_id)

        # Si es OWNER, se queda en la lista, donde puede crear más sucursales
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        profile = getattr(self.request.user, "profile", None)
        context["can_delete_branch"] = bool(profile and profile.has_role("OWNER"))
        return context
  
    def get_queryset(self) -> QuerySet[Sucursal]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return Sucursal.objects.none()
        return (
            Sucursal.objects.filter(pk__in=branch_ids)
            .select_related("company")
            .prefetch_related(
                Prefetch(
                    "branch_islands",
                    queryset=Island.objects.prefetch_related(
                        Prefetch(
                            "machines",
                            queryset=Machine.objects.prefetch_related("nozzles"),
                        )
                    ),
                ),
                Prefetch(
                    "fuel_inventories",
                    queryset=FuelInventory.objects.order_by("code"),
                ),
                Prefetch(
                    "products",
                    queryset=BranchProduct.objects.order_by("product_type", "arrival_date"),
                ),
                Prefetch(
                    "staff",
                    queryset=SucursalStaff.objects.select_related(
                        "profile__user_FK", "profile__position_FK"
                    ),
                ),
                Prefetch(
                    "shifts",
                    queryset=Shift.objects.select_related(
                        "manager__user_FK", "manager__position_FK"
                    ),
                ),
            )
        )


class SucursalCreateView(OwnerCompanyMixin, CreateView):
    form_class = SucursalForm
    template_name = "pages/sucursales/sucursal_form.html"
    success_url = reverse_lazy("sucursal_list")

    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        kwargs["company"] = self.get_company()
        return kwargs

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        form = context.get("form")
        if form is not None and "object" not in context:
            # Para la vista de creación, usamos la instancia del form
            context["object"] = form.instance
        context["can_edit_branch"] = True
        profile = getattr(self.request.user, "profile", None)
        context["can_assign_admin"] = bool(profile and profile.has_role("OWNER"))
        context.setdefault("islands", [])
        context.setdefault("fuel_inventories", [])
        return context
        

class SucursalUpdateView(OwnerCompanyMixin, UpdateView):
    model = Sucursal
    form_class = SucursalForm
    template_name = "pages/sucursales/sucursal_form.html"
    success_url = reverse_lazy("sucursal_list")
    allowed_roles = [
        "OWNER",
        "ADMINISTRATOR",
        "ACCOUNTANT",
        "ATTENDANT",
        "HEAD_ATTENDANT",
    ]
    # Edits are still blocked server-side in `post()` for non-owner/admin users via `can_edit_branch`.

    def get_queryset(self) -> QuerySet[Sucursal]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return Sucursal.objects.none()
        return (
            Sucursal.objects.filter(pk__in=branch_ids)
            .select_related("company")
            .prefetch_related(
                Prefetch(
                    "branch_islands",
                    queryset=Island.objects.order_by("number").prefetch_related(
                        Prefetch(
                            "machines",
                            queryset=Machine.objects.order_by("number").prefetch_related(
                                "nozzles"
                            ),
                        )
                    ),
                ),
                Prefetch(
                    "fuel_inventories",
                    queryset=FuelInventory.objects.order_by("code"),
                ),
                Prefetch(
                    "products",
                    queryset=BranchProduct.objects.order_by("product_type", "arrival_date"),
                ),
                Prefetch(
                    "staff",
                    queryset=SucursalStaff.objects.select_related(
                        "profile__user_FK", "profile__position_FK"
                    ),
                ),
                Prefetch(
                    "shifts",
                    queryset=Shift.objects.order_by("start_time")
                    .select_related("manager__user_FK", "manager__position_FK")
                    .prefetch_related("attendants__user_FK", "attendants__position_FK"),
                ),
            )
        )
    
    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        company = self.get_company()
        if company is None and getattr(self, "object", None):
            company = self.object.company
        kwargs["company"] = company
        return kwargs

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["database_iframe_url"] = getattr(settings, "DATABASE_IFRAME_URL", "")
        if self.object:
            islands = list(self.object.branch_islands.all())
            context["islands"] = islands
            context["island_create_form"] = IslandForm(
                initial={"sucursal": self.object}, auto_id="new-island_%s"
            )
            context["island_create_url"] = reverse(
                "sucursal_island_create", args=[self.object.pk]
            )
            shift_queryset = (
                self.object.shifts.select_related(
                    "manager__user_FK", "manager__position_FK"
                )
                .prefetch_related("attendants__user_FK", "attendants__position_FK")
                .order_by("start_time")
            )
            shifts = list(shift_queryset)
            for shift in shifts:
                shift.update_form = ShiftForm(
                    instance=shift,
                    sucursal=self.object,
                    auto_id=f"edit-shift-{shift.pk}_%s",
                )
            context["shifts"] = shifts
            profile = getattr(self.request.user, "profile", None)
            # Only owners and administrators can edit the branch
            can_edit_branch = bool(
                (profile and profile.has_role("ADMINISTRATOR"))
                or (profile and profile.has_role("OWNER"))
            )
            can_assign_admin = bool(profile and profile.has_role("OWNER"))
            can_manage_shifts = can_edit_branch
            context["can_manage_shifts"] = can_manage_shifts
            if can_manage_shifts:
                context["shift_create_form"] = ShiftForm(
                    initial={"sucursal": self.object},
                    sucursal=self.object,
                    auto_id="new-shift_%s",
                )
                context["shift_create_url"] = reverse(
                    "sucursal_shift_create", args=[self.object.pk]
                )
            if can_assign_admin:
                context["user_link_form"] = kwargs.get(
                    "user_link_form",
                    BranchUserLinkForm(
                        branch=self.object,
                        company=getattr(self.object, "company", None),
                    ),
                )
            fuel_inventories = list(self.object.fuel_inventories.all())
            for inventory in fuel_inventories:
                inventory.update_form = FuelInventoryForm(
                    instance=inventory, auto_id=f"edit-inventory-{inventory.pk}_%s"
                )
            context["fuel_inventories"] = fuel_inventories
            fuel_types = sorted({inventory.fuel_type for inventory in fuel_inventories})
            fuel_price_forms: dict[str, FuelPriceForm] = {}
            fuel_price_entries: list[dict[str, Any]] = []
            for fuel_type in fuel_types:
                modal_id = f"fuel-price-{slugify(fuel_type)}"
                latest_price = (
                    FuelPrice.objects.filter(
                        sucursal=self.object, fuel_type=fuel_type
                    )
                    .order_by("-created_at", "-pk")
                    .first()
                )
                fuel_price_entries.append(
                    {
                        "fuel_type": fuel_type,
                        "current_price": getattr(latest_price, "price", None),
                        "last_updated": getattr(latest_price, "created_at", None),
                        "modal_id": modal_id,
                    }
                )
                fuel_price_forms[fuel_type] = FuelPriceForm(
                    initial={"fuel_type": fuel_type},
                    branch=self.object,
                    available_fuel_types=fuel_types,
                    auto_id=f"{modal_id}-form_%s",
                )
                fuel_price_entries[-1]["form"] = fuel_price_forms[fuel_type]
            context["fuel_price_entries"] = fuel_price_entries
            context["fuel_price_forms"] = fuel_price_forms
            context["fuel_inventory_create_form"] = FuelInventoryForm(
                initial={"sucursal": self.object}, auto_id="new-inventory_%s"
            )
            context["fuel_inventory_create_url"] = reverse(
                "sucursal_fuel_inventory_create", args=[self.object.pk]
            )
            products = list(self.object.products.all())
            for product in products:
                product.update_form = BranchProductForm(
                    instance=product, auto_id=f"edit-product-{product.pk}_%s"
                )
            context["products"] = products
            context["product_create_form"] = BranchProductForm(
                initial={"sucursal": self.object}, auto_id="new-product_%s"
            )
            context["product_create_url"] = reverse(
                "sucursal_product_create", args=[self.object.pk]
            )
            branch_credit_sales = (
                ServiceSessionCreditSale.objects.filter(
                    service_session__shift__sucursal=self.object
                )
                .select_related(
                    "service_session__shift",
                    "responsible__user_FK",
                    "fuel_inventory",
                )
                .order_by("-created_at")
            )
            context["branch_credit_sales"] = branch_credit_sales
            context["branch_credit_sales_count"] = branch_credit_sales.count()
            context["branch_credit_sales_total"] = (
                branch_credit_sales.aggregate(
                    total=Coalesce(
                        Sum("amount"),
                        Value(0),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                )["total"]
                or 0
            )
            # --- filtros desde la URL (GET) para el historial ---
            year = self.request.GET.get("year") or ""
            month = self.request.GET.get("month") or ""
            shift_query = self.request.GET.get("shift") or ""

            # --- queryset base de sesiones cerradas de esta sucursal ---
            closed_service_sessions_qs = (
                ServiceSession.objects.filter(
                    shift__sucursal=self.object,
                    ended_at__isnull=False,
                )
                .select_related("shift__manager__user_FK")
                .prefetch_related(
                    "attendants__user_FK",
                    Prefetch(
                        "credit_sales",
                        queryset=ServiceSessionCreditSale.objects.select_related(
                            "responsible__user_FK", "fuel_inventory"
                        ),
                    ),
                    Prefetch(
                        "fuel_loads",
                        queryset=ServiceSessionFuelLoad.objects.select_related(
                            "responsible__user_FK", "inventory"
                        ),
                    ),
                    Prefetch(
                        "product_loads",
                        queryset=ServiceSessionProductLoad.objects.select_related(
                            "responsible__user_FK", "product"
                        ),
                    ),
                    Prefetch(
                        "product_sales",
                        queryset=ServiceSessionProductSale.objects.select_related(
                            "responsible__user_FK"
                        ).prefetch_related("items__product"),
                    ),
                    Prefetch(
                        "withdrawals",
                        queryset=ServiceSessionWithdrawal.objects.select_related(
                            "responsible__user_FK"
                        ),
                    ),
                    Prefetch(
                        "transbank_vouchers",
                        queryset=ServiceSessionTransbankVoucher.objects.select_related(
                            "responsible__user_FK"
                        ),
                    ),
                    Prefetch(
                        "firefighter_payments",
                        queryset=ServiceSessionFirefighterPayment.objects.select_related(
                            "firefighter__user_FK"
                        ),
                    ),
                )
                .order_by("-ended_at")
            )

            # --- aplicar filtros sobre el queryset base ---
            filtered_sessions = closed_service_sessions_qs

            if year.isdigit():
                filtered_sessions = filtered_sessions.filter(
                    ended_at__year=int(year)
                )

            if month.isdigit():
                filtered_sessions = filtered_sessions.filter(
                    ended_at__month=int(month)
                )

            if shift_query:
                filtered_sessions = filtered_sessions.filter(
                    shift__code__icontains=shift_query
                )

            # --- construir registros de historial con el queryset filtrado ---
            history_records = []
            decimal_zero = Decimal("0")
            flow_mismatch_labels = dict(ServiceSession.FLOW_MISMATCH_CHOICES)

            for session in filtered_sessions:
                credit_sales = list(session.credit_sales.all())
                fuel_loads = list(session.fuel_loads.all())
                product_loads = list(session.product_loads.all())
                product_sales = list(session.product_sales.all())
                withdrawals = list(session.withdrawals.all())
                vouchers = list(session.transbank_vouchers.all())
                firefighter_payments = list(session.firefighter_payments.all())

                product_sale_items_total = sum(
                    (
                        item.quantity
                        for sale in product_sales
                        for item in sale.items.all()
                    ),
                    0,
                )
                product_sale_value_total = sum(
                    (
                        item.quantity * item.product.value
                        for sale in product_sales
                        for item in sale.items.all()
                    ),
                    decimal_zero,
                )
                credit_total = sum(
                    ((credit.amount or decimal_zero) for credit in credit_sales),
                    decimal_zero,
                )
                fuel_load_payment_total = sum(
                    (
                        (load.payment_amount or decimal_zero)
                        for load in fuel_loads
                    ),
                    decimal_zero,
                )
                product_load_payment_total = sum(
                    (
                        (load.payment_amount or decimal_zero)
                        for load in product_loads
                    ),
                    decimal_zero,
                )
                withdrawal_total = sum(
                    ((withdrawal.amount or decimal_zero) for withdrawal in withdrawals),
                    decimal_zero,
                )
                voucher_total = sum(
                    ((voucher.total_amount or decimal_zero) for voucher in vouchers),
                    decimal_zero,
                )
                firefighter_payments_total = sum(
                    (
                        (payment.amount or decimal_zero)
                        for payment in firefighter_payments
                    ),
                    decimal_zero,
                )

                turn_profit = (
                    credit_total
                    + voucher_total
                    + withdrawal_total
                    + product_sale_value_total
                )
                net_turn_profit = (
                    turn_profit
                    - fuel_load_payment_total
                    - firefighter_payments_total
                    - product_load_payment_total
                )
                history_records.append(
                    {
                        "session": session,
                        "shift_schedule": f"{session.shift.start_time:%H:%M} - {session.shift.end_time:%H:%M}",
                        "attendants": list(session.attendants.all()),
                        "credit_sales": credit_sales,
                        "credit_count": len(credit_sales),
                        "credit_total": credit_total,
                        "withdrawals": withdrawals,
                        "fuel_load_count": len(fuel_loads),
                        "fuel_loads": fuel_loads,
                        "fuel_load_liters": sum(
                            ((load.liters_added or decimal_zero) for load in fuel_loads),
                            decimal_zero,
                        ),
                        "fuel_load_payment_total": fuel_load_payment_total,
                        "product_loads": product_loads,
                        "product_load_count": len(product_loads),
                        "product_load_quantity": sum(
                            (load.quantity_added or 0) for load in product_loads
                        ),
                        "product_load_payment_total": product_load_payment_total,
                        "product_sales": product_sales,
                        "product_sales_count": len(product_sales),
                        "product_sales_items": product_sale_items_total,
                        "product_sales_value": product_sale_value_total,
                        "vouchers": vouchers,
                        "withdrawal_total": withdrawal_total,
                        "voucher_total": voucher_total,
                        "firefighter_payments": firefighter_payments,
                        "firefighter_payments_total": firefighter_payments_total,
                        "flow_mismatch_amount": session.flow_mismatch_amount,
                        "flow_mismatch_label": flow_mismatch_labels.get(
                            session.flow_mismatch_type,
                            flow_mismatch_labels[ServiceSession.FLOW_MISMATCH_NONE],
                        ),
                        "turn_profit": turn_profit,
                        "net_turn_profit": net_turn_profit,
                    }
                )

            paginator = Paginator(history_records, 5)
            page_number = self.request.GET.get("history_page") or 1
            service_history_page = paginator.get_page(page_number)

            query_params = self.request.GET.copy()
            query_params.pop("history_page", None)

            context["service_history_page"] = service_history_page
            context["service_history"] = service_history_page.object_list
            context["service_history_total"] = paginator.count
            context["history_querystring"] = query_params.urlencode()

            # --- datos auxiliares para los selects del filtro ---
            years_qs = closed_service_sessions_qs.datetimes(
                "ended_at", "year", order="DESC"
            )
            context["history_years"] = [d.year for d in years_qs]

            context["history_months"] = [
                {"value": i, "label": calendar.month_name[i].capitalize()}
                for i in range(1, 13)
            ]

            context["history_filters"] = {
                "year": year,
                "month": month,
                "shift": shift_query,
            }

            for island in islands:
                island.update_form = IslandForm(
                    instance=island, auto_id=f"edit-island-{island.pk}_%s"
                )
                island.machine_create_form = MachineForm(
                    initial={"island": island}, auto_id=f"new-machine-{island.pk}_%s"
                )
                machines = list(island.machines.all())
                for machine in machines:
                    machine.update_form = MachineForm(
                        instance=machine, auto_id=f"edit-machine-{machine.pk}_%s"
                    )
                    if MachineFuelInventoryNumeral.objects.filter(
                        machine=machine, numeral__gt=0
                    ).exists():
                        machine.nozzle_create_form = NozzleForm(
                            auto_id=f"new-nozzle-{machine.pk}_%s",
                            initial={"machine": machine},
                            machine=machine,
                        )
                    else:
                        machine.nozzle_create_form = None
                    nozzles = list(machine.nozzles.all())
                    for nozzle in nozzles:
                        nozzle.update_form = NozzleForm(
                            machine=machine,
                            instance=nozzle, auto_id=f"edit-nozzle-{nozzle.pk}_%s"
                        )
                    machine.nozzles_list = nozzles
                    inventories = machine.get_fuel_inventories()
                    current_numerals = []
                    for inventory in inventories:
                        current_numerals.extend(
                            machine.get_numerals_for_inventory(inventory)
                        )
                    machine.current_numerals = current_numerals
                island.machines_list = machines
        else:
            context.setdefault("islands", [])
            context.setdefault("shifts", [])
            context.setdefault("products", [])
            context.setdefault("fuel_price_entries", [])
            context.setdefault("fuel_price_forms", {})
            context.setdefault("branch_credit_sales", [])
            context.setdefault("branch_credit_sales_count", 0)
            context.setdefault("branch_credit_sales_total", 0)
            context.setdefault("service_history", [])
            context.setdefault("service_history_page", None)
            context.setdefault("service_history_total", 0)
            context.setdefault("history_querystring", "")
            context.setdefault("can_manage_shifts", False)
        # Expose read/edit capability flag to template
        context["can_edit_branch"] = locals().get("can_edit_branch", False)
        context["can_assign_admin"] = locals().get("can_assign_admin", False)
        if not context.get("active_modal"):
            requested_modal = self.request.GET.get("modal")
            if requested_modal:
                context["active_modal"] = requested_modal
        context.setdefault("active_modal", None)
        return context

    def get_success_url(self) -> str:
        if self.object:
            return reverse("sucursal_update", args=[self.object.pk])
        return super().get_success_url()

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        profile = getattr(request.user, "profile", None)
        can_edit_branch = bool(
            (profile and profile.has_role("ADMINISTRATOR"))
            or (profile and profile.has_role("OWNER"))
        )
        if request.method == "POST" and not can_edit_branch:
            messages.error(request, "No tienes permisos para editar esta sucursal.")
            return redirect("sucursal_update", pk=self.object.pk)
        scope = request.POST.get("form_scope")

        handler_map = {
            "shift-update": "_handle_shift_update",
            "fuel-inventory-update": "_handle_fuel_inventory_update",
            "fuel-price-create": "_handle_fuel_price_create",
            "product-update": "_handle_product_update",
            "island-update": "_handle_island_update",
            "machine-update": "_handle_machine_update",
            "nozzle-update": "_handle_nozzle_update",
            "branch-user-link": "_handle_branch_user_link",
        }

        handler_name = handler_map.get(scope)
        if handler_name:
            handler = getattr(self, handler_name, None)
            if handler:
                return handler(request)
            messages.error(request, "No se pudo procesar el formulario solicitado.")
            return redirect("sucursal_update", pk=self.object.pk)

        return super().post(request, *args, **kwargs)

    def _get_branch_form(self) -> SucursalForm:
        form_kwargs = self.get_form_kwargs()
        form_kwargs.update({"data": None, "files": None, "instance": self.object})
        return self.form_class(**form_kwargs)

    def _render_with_inline_form(self, *, modal_name: str) -> Any:
        context = self.get_context_data(form=self._get_branch_form())
        context["active_modal"] = modal_name
        return self.render_to_response(context)

    def _handle_branch_user_link(self, request) -> Any:
        profile = getattr(request.user, "profile", None)
        if not profile or not profile.has_role("OWNER"):
            messages.error(
                request,
                "No tienes permisos para asignar administradores a esta sucursal.",
            )
            return redirect("sucursal_update", pk=self.object.pk)

        form = BranchUserLinkForm(
            request.POST,
            branch=self.object,
            company=getattr(self.object, "company", None),
        )
        if form.is_valid():
            form.save()
            messages.success(
                request, "Usuarios asignados a la sucursal actualizados correctamente."
            )
            return redirect("sucursal_update", pk=self.object.pk)

        messages.error(
            request, "No se pudo actualizar el personal de la sucursal. Revisa los datos."
        )
        context = self.get_context_data(
            form=self._get_branch_form(), user_link_form=form
        )
        return self.render_to_response(context)

    def _handle_shift_update(self, request) -> Any:
        shift_id = request.POST.get("object_id")
        shift = get_object_or_404(self.object.shifts.all(), pk=shift_id)
        form = ShiftForm(
            request.POST,
            instance=shift,
            sucursal=self.object,
            auto_id=f"edit-shift-{shift.pk}_%s",
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Turno actualizado correctamente.")
            return redirect("sucursal_update", pk=self.object.pk)
        response = self._render_with_inline_form(
            modal_name=f"shift-edit-{shift.pk}"
        )
        for shift_context in response.context_data.get("shifts", []):  # type: ignore[attr-defined]
            if shift_context.pk == shift.pk:
                shift_context.update_form = form
                break
        return response

    def _handle_fuel_inventory_update(self, request) -> Any:
        inventory_id = request.POST.get("object_id")
        inventory = get_object_or_404(
            self.object.fuel_inventories.all(), pk=inventory_id
        )
        form = FuelInventoryForm(
            request.POST,
            instance=inventory,
            auto_id=f"edit-inventory-{inventory.pk}_%s",
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Inventario actualizado correctamente.")
            return redirect("sucursal_update", pk=self.object.pk)
        response = self._render_with_inline_form(
            modal_name=f"fuel-inventory-edit-{inventory.pk}"
        )
        for inventory_context in response.context_data.get("fuel_inventories", []):  # type: ignore[attr-defined]
            if inventory_context.pk == inventory.pk:
                inventory_context.update_form = form
                break
        return response

    def _handle_fuel_price_create(self, request) -> Any:
        fuel_type = request.POST.get("fuel_type", "")
        available_types = list(
            self.object.fuel_inventories.values_list("fuel_type", flat=True).distinct()
        )
        modal_id = f"fuel-price-{slugify(fuel_type)}" if fuel_type else "fuel-price"
        form = FuelPriceForm(
            request.POST,
            branch=self.object,
            available_fuel_types=available_types,
            auto_id=f"{modal_id}-form_%s",
        )
        if form.is_valid():
            form.save()
            messages.success(
                request,
                "Precio asignado correctamente para el combustible seleccionado.",
            )
            return redirect("sucursal_update", pk=self.object.pk)

        response = self._render_with_inline_form(modal_name=modal_id)
        fuel_price_forms = response.context_data.get("fuel_price_forms", {})  # type: ignore[attr-defined]
        fuel_price_forms[fuel_type] = form
        response.context_data["fuel_price_forms"] = fuel_price_forms  # type: ignore[index]
        for entry in response.context_data.get("fuel_price_entries", []):  # type: ignore[attr-defined]
            if entry.get("fuel_type") == fuel_type:
                entry["form"] = form
                break
        return response

    def _handle_product_update(self, request) -> Any:
        product_id = request.POST.get("object_id")
        product = get_object_or_404(self.object.products.all(), pk=product_id)
        form = BranchProductForm(
            request.POST,
            instance=product,
            auto_id=f"edit-product-{product.pk}_%s",
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Producto actualizado correctamente.")
            return redirect("sucursal_update", pk=self.object.pk)
        response = self._render_with_inline_form(
            modal_name=f"product-edit-{product.pk}"
        )
        for product_context in response.context_data.get("products", []):  # type: ignore[attr-defined]
            if product_context.pk == product.pk:
                product_context.update_form = form
                break
        return response

    def _handle_island_update(self, request) -> Any:
        island_id = request.POST.get("object_id")
        island = get_object_or_404(self.object.branch_islands.all(), pk=island_id)
        form = IslandForm(
            request.POST,
            instance=island,
            auto_id=f"edit-island-{island.pk}_%s",
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Isla actualizada correctamente.")
            return redirect("sucursal_update", pk=self.object.pk)
        response = self._render_with_inline_form(
            modal_name=f"island-edit-{island.pk}"
        )
        for island_context in response.context_data.get("islands", []):  # type: ignore[attr-defined]
            if island_context.pk == island.pk:
                island_context.update_form = form
                break
        return response

    def _handle_machine_update(self, request) -> Any:
        machine_id = request.POST.get("object_id")
        machine = get_object_or_404(
            Machine.objects.filter(island__sucursal=self.object), pk=machine_id
        )
        form = MachineForm(
            request.POST,
            instance=machine,
            island=machine.island,
            auto_id=f"edit-machine-{machine.pk}_%s",
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Máquina actualizada correctamente.")
            return redirect("sucursal_update", pk=self.object.pk)
        response = self._render_with_inline_form(
            modal_name=f"machine-edit-{machine.pk}"
        )
        for island_context in response.context_data.get("islands", []):  # type: ignore[attr-defined]
            machines = getattr(island_context, "machines_list", [])
            for machine_context in machines:
                if machine_context.pk == machine.pk:
                    machine_context.update_form = form
                    break
        return response

    def _handle_nozzle_update(self, request) -> Any:
        nozzle_id = request.POST.get("object_id")
        nozzle = get_object_or_404(
            Nozzle.objects.filter(machine__island__sucursal=self.object), pk=nozzle_id
        )
        form = NozzleForm(
            request.POST,
            instance=nozzle,
            machine=nozzle.machine,
            auto_id=f"edit-nozzle-{nozzle.pk}_%s",
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Pistola actualizada correctamente.")
            return redirect("sucursal_update", pk=self.object.pk)
        response = self._render_with_inline_form(
            modal_name=f"nozzle-edit-{nozzle.pk}"
        )
        for island_context in response.context_data.get("islands", []):  # type: ignore[attr-defined]
            machines = getattr(island_context, "machines_list", [])
            for machine_context in machines:
                nozzles = getattr(machine_context, "nozzles_list", [])
                for nozzle_context in nozzles:
                    if nozzle_context.pk == nozzle.pk:
                        nozzle_context.update_form = form
                        break
        return response


class ServiceSessionSummaryExportView(OwnerCompanyMixin, View):
    """
    Exporta el resumen de un servicio (turno) a un XLSX abrible en Excel.
    """

    # Roles que pueden exportar
    allowed_roles = ["OWNER", "ADMINISTRATOR", "ACCOUNTANT", "HEAD_ATTENDANT"]

    def get(self, request, branch_pk, pk, *args, **kwargs):
        profile = request.user.profile

        # Validar que la sucursal está dentro del alcance del usuario
        managed_branch_ids = set(self.get_managed_branch_ids())
        if managed_branch_ids and branch_pk not in managed_branch_ids:
            raise PermissionDenied("No tienes permisos para esta sucursal.")

        # Obtenemos la sesión del servicio SOLO de esa sucursal
        session = get_object_or_404(
            ServiceSession.objects.select_related(
                "shift__sucursal",
                "shift__manager__user_FK",
            ).prefetch_related(
                "attendants__user_FK",
                "credit_sales",
                "fuel_loads",
                "product_loads",
                "product_sales__items__product",
                "withdrawals",
                "transbank_vouchers",
                "firefighter_payments",
            ),
            pk=pk,
            shift__sucursal_id=branch_pk,
        )

        decimal_zero = Decimal("0")

        credit_sales = list(session.credit_sales.all())
        fuel_loads = list(session.fuel_loads.all())
        product_loads = list(session.product_loads.all())
        product_sales = list(session.product_sales.all())
        withdrawals = list(session.withdrawals.all())
        vouchers = list(session.transbank_vouchers.all())
        firefighter_payments = list(session.firefighter_payments.all())

        # === Cálculos igual que en el historial ===
        product_sale_items_total = sum(
            (item.quantity for sale in product_sales for item in sale.items.all()),
            0,
        )
        product_sale_value_total = sum(
            (
                item.quantity * item.product.value
                for sale in product_sales
                for item in sale.items.all()
            ),
            decimal_zero,
        )

        credit_total = sum(
            ((credit.amount or decimal_zero) for credit in credit_sales),
            decimal_zero,
        )

        fuel_load_payment_total = sum(
            ((load.payment_amount or decimal_zero) for load in fuel_loads),
            decimal_zero,
        )

        product_load_payment_total = sum(
            ((load.payment_amount or decimal_zero) for load in product_loads),
            decimal_zero,
        )

        withdrawal_total = sum(
            ((withdrawal.amount or decimal_zero) for withdrawal in withdrawals),
            decimal_zero,
        )

        voucher_total = sum(
            ((voucher.total_amount or decimal_zero) for voucher in vouchers),
            decimal_zero,
        )

        firefighter_payments_total = sum(
            ((payment.amount or decimal_zero) for payment in firefighter_payments),
            decimal_zero,
        )

        turn_profit = (
            credit_total
            + voucher_total
            + withdrawal_total
            + product_sale_value_total
        )

        net_turn_profit = (
            turn_profit
            - fuel_load_payment_total
            - firefighter_payments_total
            - product_load_payment_total
        )

        # ================== GENERAR XLSX ==================
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Turno"

        headers = [
            "FECHA",
            "TURNO",
            "TOTAL VENTA",
            "CREDITO",
            "CANT.VOUCHER",
            "MONTO VOUCHER",
            "CANT.TIRADA",
            "MONTO TIRADA",
            "PAGOS BOMBEROS",
            "GASTO STOCK",
            "PAGO COMBUSTIBLE",
            "GANANCIA TURNO",
            "GANANCIA REAL",
        ]

        ws.append(headers)

        data_row = [
            session.started_at.date() if session.started_at else "",
            session.shift.code if session.shift else "",
            float(product_sale_value_total),      # TOTAL VENTA
            float(credit_total),                  # CREDITO
            len(vouchers),                        # CANT.VOUCHER
            float(voucher_total),                 # MONTO VOUCHER
            len(withdrawals),                     # CANT.TIRADA
            float(withdrawal_total),              # MONTO TIRADA
            float(firefighter_payments_total),    # PAGOS BOMBEROS
            float(product_load_payment_total),    # GASTO STOCK
            float(fuel_load_payment_total),       # PAGO COMBUSTIBLE
            float(turn_profit),                   # GANANCIA TURNO
            float(net_turn_profit),               # GANANCIA REAL
        ]

        ws.append(data_row)

        # Estilos básicos: negrita en header y auto-ancho
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = max(len(header) + 2, 14)

        # Formato numérico con separador de miles
        money_cols = range(3, 14)  # columnas C a M
        for col_idx in money_cols:
            cell = ws.cell(row=2, column=col_idx)
            cell.number_format = "#,##0"

        # Exportar a memoria y responder
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"resumen_turno_{session.shift.code}_{session.pk}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        return response

class ServiceHistoryExportView(OwnerCompanyMixin, View):
    """
    Exporta un XLSX con todos los servicios cerrados de una sucursal,
    aplicando los mismos filtros del Historial (año, mes, turno).
    """

    allowed_roles = ["OWNER", "ADMINISTRATOR", "ACCOUNTANT", "HEAD_ATTENDANT"]

    def get(self, request, branch_pk, *args, **kwargs):
        from sucursalApp.models import Sucursal, ServiceSession
        from sucursalApp.models import (
            ServiceSessionCreditSale,
            ServiceSessionFuelLoad,
            ServiceSessionProductLoad,
            ServiceSessionProductSale,
            ServiceSessionWithdrawal,
            ServiceSessionTransbankVoucher,
            ServiceSessionFirefighterPayment,
        )
        from django.db.models import Prefetch

        branch_ids = self.get_managed_branch_ids()
        branch = get_object_or_404(
            Sucursal.objects.filter(pk__in=branch_ids), pk=branch_pk
        )

        # --- filtros desde la URL, mismos nombres que en el Historial ---
        year = request.GET.get("year") or ""
        month = request.GET.get("month") or ""
        shift_query = request.GET.get("shift") or ""

        # --- queryset base: servicios cerrados de la sucursal ---
        sessions_qs = (
            ServiceSession.objects.filter(
                shift__sucursal=branch,
                ended_at__isnull=False,
            )
            .select_related("shift__sucursal", "shift__manager__user_FK")
            .prefetch_related(
                "attendants__user_FK",
                Prefetch(
                    "credit_sales",
                    queryset=ServiceSessionCreditSale.objects.select_related(
                        "responsible__user_FK", "fuel_inventory"
                    ),
                ),
                Prefetch(
                    "fuel_loads",
                    queryset=ServiceSessionFuelLoad.objects.select_related(
                        "responsible__user_FK", "inventory"
                    ),
                ),
                Prefetch(
                    "product_loads",
                    queryset=ServiceSessionProductLoad.objects.select_related(
                        "responsible__user_FK", "product"
                    ),
                ),
                Prefetch(
                    "product_sales",
                    queryset=ServiceSessionProductSale.objects.select_related(
                        "responsible__user_FK"
                    ).prefetch_related("items__product"),
                ),
                Prefetch(
                    "withdrawals",
                    queryset=ServiceSessionWithdrawal.objects.select_related(
                        "responsible__user_FK"
                    ),
                ),
                Prefetch(
                    "transbank_vouchers",
                    queryset=ServiceSessionTransbankVoucher.objects.select_related(
                        "responsible__user_FK"
                    ),
                ),
                Prefetch(
                    "firefighter_payments",
                    queryset=ServiceSessionFirefighterPayment.objects.select_related(
                        "firefighter__user_FK"
                    ),
                ),
            )
            .order_by("-ended_at")
        )

        # --- aplicar filtros (mismos que en get_context_data del Historial) ---
        if year.isdigit():
            sessions_qs = sessions_qs.filter(ended_at__year=int(year))

        if month.isdigit():
            sessions_qs = sessions_qs.filter(ended_at__month=int(month))

        if shift_query:
            sessions_qs = sessions_qs.filter(shift__code__icontains=shift_query)

        # --- crear workbook ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial servicios"

        headers = [
            "FECHA",
            "TURNO",
            "TOTAL VENTA",
            "CRÉDITO",
            "CANT. VOUCHERS",
            "MONTO VOUCHERS",
            "CANT. TIRADAS",
            "MONTO TIRADAS",
            "PAGOS BOMBEROS",
            "GASTO STOCK",
            "PAGO COMBUSTIBLE",
            "GANANCIA TURNO",
            "GANANCIA REAL",
        ]

        ws.append(headers)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 14)

        decimal_zero = Decimal("0")

        # --- recorrer sesiones y calcular totales (igual que en Historial) ---
        row = 2
        for session in sessions_qs:
            credit_sales = list(session.credit_sales.all())
            fuel_loads = list(session.fuel_loads.all())
            product_loads = list(session.product_loads.all())
            product_sales = list(session.product_sales.all())
            withdrawals = list(session.withdrawals.all())
            vouchers = list(session.transbank_vouchers.all())
            firefighter_payments = list(session.firefighter_payments.all())

            product_sale_value_total = sum(
                (
                    item.quantity * item.product.value
                    for sale in product_sales
                    for item in sale.items.all()
                ),
                decimal_zero,
            )
            credit_total = sum(
                ((c.amount or decimal_zero) for c in credit_sales),
                decimal_zero,
            )
            fuel_load_payment_total = sum(
                ((l.payment_amount or decimal_zero) for l in fuel_loads),
                decimal_zero,
            )
            product_load_payment_total = sum(
                ((l.payment_amount or decimal_zero) for l in product_loads),
                decimal_zero,
            )
            withdrawal_total = sum(
                ((w.amount or decimal_zero) for w in withdrawals),
                decimal_zero,
            )
            voucher_total = sum(
                ((v.total_amount or decimal_zero) for v in vouchers),
                decimal_zero,
            )
            firefighter_payments_total = sum(
                ((p.amount or decimal_zero) for p in firefighter_payments),
                decimal_zero,
            )

            turn_profit = (
                (session.initial_budget or decimal_zero)
                + credit_total
                + voucher_total
                + withdrawal_total
                + product_sale_value_total
            )
            net_turn_profit = (
                turn_profit
                - fuel_load_payment_total
                - firefighter_payments_total
                - product_load_payment_total
            )

            ws.cell(row=row, column=1, value=session.started_at.date())
            ws.cell(row=row, column=2, value=session.shift.code)
            ws.cell(row=row, column=3, value=float(product_sale_value_total))
            ws.cell(row=row, column=4, value=float(credit_total))
            ws.cell(row=row, column=5, value=len(vouchers))
            ws.cell(row=row, column=6, value=float(voucher_total))
            ws.cell(row=row, column=7, value=len(withdrawals))
            ws.cell(row=row, column=8, value=float(withdrawal_total))
            ws.cell(row=row, column=9, value=float(firefighter_payments_total))
            ws.cell(row=row, column=10, value=float(product_load_payment_total))
            ws.cell(row=row, column=11, value=float(fuel_load_payment_total))
            ws.cell(row=row, column=12, value=float(turn_profit))
            ws.cell(row=row, column=13, value=float(net_turn_profit))

            row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"historial_servicios_{branch.name}_{branch_pk}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class BranchStaffManageView(OwnerCompanyMixin, SingleObjectMixin, FormView):
    model = Sucursal
    form_class = BranchStaffForm
    template_name = "pages/usuarios/usuarios_lista.html"
    allowed_roles = ["OWNER", "ADMINISTRATOR"]
    http_method_names = ["post"]

    def get_queryset(self):
        return self.get_managed_branches_queryset()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        branch = self.get_object()
        profile = getattr(self.request.user, "profile", None)
        kwargs.update(
            {
                "instance": branch,
                "company": getattr(branch, "company", None),
                "allow_admin_assignment": bool(profile and profile.is_owner()),
            }
        )
        return kwargs

    def form_valid(self, form):
        form.save()
        messages.success(
            self.request, "Personal de la sucursal actualizado correctamente."
        )
        return redirect("User")

    def form_invalid(self, form):
        messages.error(
            self.request,
            "No se pudo actualizar el personal de la sucursal. Revisa los datos.",
        )
        return redirect("User")

class SucursalDeleteView(OwnerCompanyMixin, DeleteView):
    model = Sucursal
    success_url = reverse_lazy("sucursal_list")

    def get_queryset(self) -> QuerySet[Sucursal]:
        company = self.get_company()
        if company is None:
            return Sucursal.objects.none()
        return Sucursal.objects.filter(company=company)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        sucursal_name = self.object.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f"Sucursal '{sucursal_name}' eliminada con éxito.")
        return response

class BranchAccessMixin(OwnerCompanyMixin):
    branch_url_kwarg = "branch_pk"
    allowed_roles = ["OWNER", "ADMINISTRATOR"]

    def get_branch_queryset(self):
        return self.get_managed_branches_queryset()

    def get_sucursal(self) -> Sucursal:
        queryset = self.get_branch_queryset()
        return get_object_or_404(queryset, pk=self.kwargs.get(self.branch_url_kwarg))
    def dispatch(self, request, *args, **kwargs):
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            return redirect(self.redirect_url)

class ShiftAccessMixin(OwnerCompanyMixin):
    model = Shift
    allowed_roles = ["OWNER", "ADMINISTRATOR"]
    def get_queryset(self) -> QuerySet[Shift]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return Shift.objects.none()
        return (
            Shift.objects.filter(sucursal_id__in=branch_ids)
            .select_related("sucursal", "manager__user_FK", "manager__position_FK")
            .prefetch_related("attendants__user_FK", "attendants__position_FK")
        )

    def get_object(self) -> Shift:
        return get_object_or_404(self.get_queryset(), pk=self.kwargs.get("pk"))

    def get_success_url(self, obj: Shift | None = None) -> str:
        instance = obj or getattr(self, "object", None)
        if instance is None:
            instance = self.get_object()
        return reverse("sucursal_update", args=[instance.sucursal_id])


class ShiftCreateView(BranchAccessMixin, CreateView):
    form_class = ShiftForm
    template_name = "pages/sucursales/related_form.html"
    allowed_roles = ["ADMINISTRATOR","OWNER"]

    def get_initial(self) -> Dict[str, Any]:
        initial = super().get_initial()
        initial.setdefault("sucursal", self.get_sucursal())
        return initial

    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        kwargs.setdefault("sucursal", self.get_sucursal())
        return kwargs

    def form_valid(self, form: ShiftForm) -> HttpResponseRedirect:
        form.instance.sucursal = self.get_sucursal()
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        sucursal = self.get_sucursal()
        context.update({"sucursal": sucursal, "title": "Agregar turno"})
        return context

    def get_success_url(self) -> str:
        return reverse("sucursal_update", args=[self.get_sucursal().pk])


class ShiftUpdateView(ShiftAccessMixin, UpdateView):
    form_class = ShiftForm
    template_name = "pages/sucursales/related_form.html"
    allowed_roles = ["ADMINISTRATOR", "OWNER"]

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return redirect_to_modal(
            self.object.sucursal_id, f"shift-edit-{self.object.pk}"
        )

    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        kwargs.setdefault("sucursal", self.object.sucursal)
        return kwargs

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update({"sucursal": self.object.sucursal, "title": "Editar turno"})
        return context

    def get_success_url(self) -> str:
        return super().get_success_url(self.object)


class ShiftDeleteView(ShiftAccessMixin, View):
    allowed_roles = ["ADMINISTRATOR","OWNER"]
    def post(self, request, *args, **kwargs) -> HttpResponseRedirect:
        shift = self.get_object()
        success_url = self.get_success_url(shift)
        shift.delete()
        return HttpResponseRedirect(success_url)


class FuelInventoryAccessMixin(OwnerCompanyMixin):
    model = FuelInventory
    allowed_roles = ["OWNER", "ADMINISTRATOR"]
    def get_queryset(self) -> QuerySet[FuelInventory]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return FuelInventory.objects.none()
        return FuelInventory.objects.filter(sucursal_id__in=branch_ids).select_related(
            "sucursal"
        )

    def get_object(self) -> FuelInventory:
        return get_object_or_404(self.get_queryset(), pk=self.kwargs.get("pk"))

    def get_success_url(self, obj: FuelInventory | None = None) -> str:
        instance = obj or getattr(self, "object", None)
        if instance is None:
            instance = self.get_object()
        return reverse("sucursal_update", args=[instance.sucursal_id])


class FuelInventoryCreateView(BranchAccessMixin, CreateView):
    form_class = FuelInventoryForm
    template_name = "pages/sucursales/related_form.html"

    def get_initial(self) -> Dict[str, Any]:
        initial = super().get_initial()
        initial.setdefault("sucursal", self.get_sucursal())
        return initial

    def form_valid(self, form: FuelInventoryForm) -> HttpResponseRedirect:
        form.instance.sucursal = self.get_sucursal()
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        sucursal = self.get_sucursal()
        context.update({"sucursal": sucursal, "title": "Agregar inventario"})
        return context

    def get_success_url(self) -> str:
        return reverse("sucursal_update", args=[self.get_sucursal().pk])


class FuelInventoryUpdateView(FuelInventoryAccessMixin, UpdateView):
    form_class = FuelInventoryForm
    template_name = "pages/sucursales/related_form.html"

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return redirect_to_modal(
            self.object.sucursal_id, f"fuel-inventory-edit-{self.object.pk}"
        )

    def form_valid(self, form: FuelInventoryForm) -> HttpResponseRedirect:
        form.instance.sucursal = self.object.sucursal
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "sucursal": self.object.sucursal,
                "title": "Editar inventario",
            }
        )
        return context

    def get_success_url(self) -> str:
        return super().get_success_url(self.object)


class FuelInventoryDeleteView(FuelInventoryAccessMixin, View):
    def post(self, request, *args, **kwargs) -> HttpResponseRedirect:
        inventory = self.get_object()
        success_url = self.get_success_url(inventory)
        inventory.delete()
        return HttpResponseRedirect(success_url)


class BranchProductAccessMixin(OwnerCompanyMixin):
    model = BranchProduct
    allowed_roles = ["OWNER", "ADMINISTRATOR"]

    def get_queryset(self) -> QuerySet[BranchProduct]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return BranchProduct.objects.none()
        return BranchProduct.objects.filter(sucursal_id__in=branch_ids).select_related(
            "sucursal"
        )

    def get_object(self) -> BranchProduct:
        return get_object_or_404(self.get_queryset(), pk=self.kwargs.get("pk"))

    def get_success_url(self, obj: BranchProduct | None = None) -> str:
        instance = obj or getattr(self, "object", None)
        if instance is None:
            instance = self.get_object()
        return reverse("sucursal_update", args=[instance.sucursal_id])


class BranchProductCreateView(BranchAccessMixin, CreateView):
    form_class = BranchProductForm
    template_name = "pages/sucursales/related_form.html"

    def get_initial(self) -> Dict[str, Any]:
        initial = super().get_initial()
        initial.setdefault("sucursal", self.get_sucursal())
        return initial

    def form_valid(self, form: BranchProductForm) -> HttpResponseRedirect:
        form.instance.sucursal = self.get_sucursal()
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update({"sucursal": self.get_sucursal(), "title": "Agregar producto"})
        return context

    def get_success_url(self) -> str:
        return reverse("sucursal_update", args=[self.get_sucursal().pk])


class BranchProductUpdateView(BranchProductAccessMixin, UpdateView):
    form_class = BranchProductForm
    template_name = "pages/sucursales/related_form.html"


    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return redirect_to_modal(
            self.object.sucursal_id, f"product-edit-{self.object.pk}"
        )

    def form_valid(self, form: BranchProductForm) -> HttpResponseRedirect:
        form.instance.sucursal = self.object.sucursal
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "sucursal": self.object.sucursal,
                "title": "Editar producto",
            }
        )
        return context

    def get_success_url(self) -> str:
        return super().get_success_url(self.object)


class BranchProductDeleteView(BranchProductAccessMixin, View):
    def post(self, request, *args, **kwargs) -> HttpResponseRedirect:
        product = self.get_object()
        success_url = self.get_success_url(product)
        product.delete()
        return HttpResponseRedirect(success_url)


class CreditSaleAccessMixin(OwnerCompanyMixin):
    model = ServiceSessionCreditSale
    allowed_roles = ["OWNER", "ADMINISTRATOR"]

    def get_queryset(self) -> QuerySet[ServiceSessionCreditSale]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return ServiceSessionCreditSale.objects.none()
        return ServiceSessionCreditSale.objects.filter(
            service_session__shift__sucursal_id__in=branch_ids
        ).select_related("service_session__shift__sucursal")

    def get_object(self) -> ServiceSessionCreditSale:
        return get_object_or_404(self.get_queryset(), pk=self.kwargs.get("pk"))

    def get_success_url(self, obj: ServiceSessionCreditSale | None = None) -> str:
        instance = obj or getattr(self, "object", None)
        if instance is None:
            instance = self.get_object()
        return reverse("sucursal_update", args=[instance.service_session.shift.sucursal_id])

class IslandAccessMixin(OwnerCompanyMixin):
    model = Island
    allowed_roles = ["OWNER", "ADMINISTRATOR"]
    def get_queryset(self) -> QuerySet[Island]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return Island.objects.none()
        return Island.objects.filter(sucursal_id__in=branch_ids).select_related(
            "sucursal"
        )

    def get_object(self) -> Island:
        return get_object_or_404(self.get_queryset(), pk=self.kwargs.get("pk"))

    def get_success_url(self, obj: Island | None = None) -> str:
        instance = obj or getattr(self, "object", None)
        if instance is None:
            instance = self.get_object()
        return reverse("sucursal_update", args=[instance.sucursal_id])


class MachineAccessMixin(OwnerCompanyMixin):
    model = Machine
    allowed_roles = ["OWNER", "ADMINISTRATOR"]
    def get_queryset(self) -> QuerySet[Machine]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return Machine.objects.none()
        return Machine.objects.filter(island__sucursal_id__in=branch_ids).select_related(
            "island__sucursal"
        )

    def get_object(self) -> Machine:
        return get_object_or_404(self.get_queryset(), pk=self.kwargs.get("pk"))

    def get_success_url(self, obj: Machine | None = None) -> str:
        instance = obj or getattr(self, "object", None)
        if instance is None:
            instance = self.get_object()
        return reverse("sucursal_update", args=[instance.island.sucursal_id])


class CreditSaleMarkPaidView(CreditSaleAccessMixin, View):
    def post(self, request, *args, **kwargs) -> HttpResponseRedirect:
        credit_sale = self.get_object()
        if credit_sale.status != ServiceSessionCreditSale.Status.PAID:
            ServiceSessionCreditSale.objects.filter(pk=credit_sale.pk).update(
                status=ServiceSessionCreditSale.Status.PAID
            )
            messages.success(request, "El crédito fue marcado como pagado.")
        else:
            messages.info(request, "El crédito ya estaba pagado.")
        return HttpResponseRedirect(self.get_success_url(credit_sale))


class CreditSaleDeleteView(CreditSaleAccessMixin, View):
    def post(self, request, *args, **kwargs) -> HttpResponseRedirect:
        credit_sale = self.get_object()
        success_url = self.get_success_url(credit_sale)
        credit_sale.delete()
        messages.success(request, "El crédito fue eliminado correctamente.")
        return HttpResponseRedirect(success_url)

class NozzleAccessMixin(OwnerCompanyMixin):
    model = Nozzle
    allowed_roles = ["OWNER", "ADMINISTRATOR"]
    def get_queryset(self) -> QuerySet[Nozzle]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return Nozzle.objects.none()
        return Nozzle.objects.filter(
            machine__island__sucursal_id__in=branch_ids
        ).select_related("machine__island__sucursal")

    def get_object(self) -> Nozzle:
        return get_object_or_404(self.get_queryset(), pk=self.kwargs.get("pk"))

    def get_success_url(self, obj: Nozzle | None = None) -> str:
        instance = obj or getattr(self, "object", None)
        if instance is None:
            instance = self.get_object()
        return reverse("sucursal_update", args=[instance.machine.island.sucursal_id])


class IslandCreateView(BranchAccessMixin, CreateView):
    form_class = IslandForm
    template_name = "pages/sucursales/related_form.html"

    def get_initial(self) -> Dict[str, Any]:
        initial = super().get_initial()
        initial.setdefault("sucursal", self.get_sucursal())
        return initial

    def form_valid(self, form: IslandForm) -> HttpResponseRedirect:
        form.instance.sucursal = self.get_sucursal()
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        sucursal = self.get_sucursal()
        context.update(
            {
                "sucursal": sucursal,
                "title": "Agregar isla",
            }
        )
        return context

    def get_success_url(self) -> str:
        return reverse("sucursal_update", args=[self.get_sucursal().pk])


class IslandUpdateView(IslandAccessMixin, UpdateView):
    form_class = IslandForm
    template_name = "pages/sucursales/related_form.html"

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return redirect_to_modal(
            self.object.sucursal_id, f"island-edit-{self.object.pk}"
        )

    def form_valid(self, form: IslandForm) -> HttpResponseRedirect:
        form.instance.sucursal = self.object.sucursal
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "sucursal": self.object.sucursal,
                "title": "Editar isla",
            }
        )
        return context

    def get_success_url(self) -> str:
        return super().get_success_url(self.object)


class IslandDeleteView(IslandAccessMixin, View):
    def post(self, request, *args, **kwargs) -> HttpResponseRedirect:
        island = self.get_object()
        success_url = self.get_success_url(island)
        island.delete()
        return HttpResponseRedirect(success_url)


class MachineCreateView(BranchAccessMixin, CreateView):
    form_class = MachineForm
    template_name = "pages/sucursales/related_form.html"
    island_url_kwarg = "island_pk"

    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        kwargs.setdefault("island", self.get_island())
        return kwargs


    def get_island(self) -> Island:
        sucursal = self.get_sucursal()
        queryset = sucursal.branch_islands.all()
        return get_object_or_404(queryset, pk=self.kwargs.get(self.island_url_kwarg))

    def get_initial(self) -> Dict[str, Any]:
        initial = super().get_initial()
        initial.setdefault("island", self.get_island())
        return initial

    def form_valid(self, form: MachineForm) -> HttpResponseRedirect:
        form.instance.island = self.get_island()
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        island = self.get_island()
        context.update(
            {
                "sucursal": island.sucursal,
                "island": island,
                "title": "Agregar máquina",
            }
        )
        return context

    def get_success_url(self) -> str:
        return reverse("sucursal_update", args=[self.get_island().sucursal_id])


class MachineUpdateView(MachineAccessMixin, UpdateView):
    form_class = MachineForm
    template_name = "pages/sucursales/related_form.html"

    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        kwargs.setdefault("island", getattr(self, "object", None) and self.object.island)
        return kwargs

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return redirect_to_modal(
            self.object.island.sucursal_id, f"machine-edit-{self.object.pk}"
        )

    def form_valid(self, form: MachineForm) -> HttpResponseRedirect:
        form.instance.island = self.object.island
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "sucursal": self.object.island.sucursal,
                "island": self.object.island,
                "title": "Editar máquina",
            }
        )
        return context

    def get_success_url(self) -> str:
        return super().get_success_url(self.object)


class MachineDeleteView(MachineAccessMixin, View):
    def post(self, request, *args, **kwargs) -> HttpResponseRedirect:
        machine = self.get_object()
        success_url = self.get_success_url(machine)
        machine.delete()
        return HttpResponseRedirect(success_url)


class NozzleCreateView(OwnerCompanyMixin, CreateView):
    form_class = NozzleForm
    template_name = "pages/sucursales/related_form.html"
    machine_url_kwarg = "machine_pk"
    allowed_roles = ["OWNER", "ADMINISTRATOR"]

    def get_machine_queryset(self) -> QuerySet[Machine]:
        branch_ids = self.get_managed_branch_ids()
        if not branch_ids:
            return Machine.objects.none()
        return Machine.objects.filter(island__sucursal_id__in=branch_ids).select_related(
            "island__sucursal"
        )

    def get_machine(self) -> Machine:
        queryset = self.get_machine_queryset()
        return get_object_or_404(queryset, pk=self.kwargs.get(self.machine_url_kwarg))

    def get_initial(self) -> Dict[str, Any]:
        initial = super().get_initial()
        initial.setdefault("machine", self.get_machine())
        return initial


    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        kwargs.setdefault("machine", getattr(self, "object", None) or self.get_machine())
        return kwargs

    def form_valid(self, form: NozzleForm) -> HttpResponseRedirect:
        form.instance.machine = self.get_machine()
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        machine = self.get_machine()
        context.update(
            {
                "sucursal": machine.island.sucursal,
                "island": machine.island,
                "machine": machine,
                "title": "Agregar pistola",
            }
        )
        return context

    def get_success_url(self) -> str:
        return reverse("sucursal_update", args=[self.get_machine().island.sucursal_id])


class NozzleUpdateView(NozzleAccessMixin, UpdateView):
    form_class = NozzleForm
    template_name = "pages/sucursales/related_form.html"

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return redirect_to_modal(
            self.object.machine.island.sucursal_id,
            f"nozzle-edit-{self.object.pk}",
        )

    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        kwargs.setdefault("machine", getattr(self, "object", None) and self.object.machine)
        return kwargs

    def form_valid(self, form: NozzleForm) -> HttpResponseRedirect:
        form.instance.machine = self.object.machine
        return super().form_valid(form)

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "sucursal": self.object.machine.island.sucursal,
                "island": self.object.machine.island,
                "machine": self.object.machine,
                "title": "Editar pistola",
            }
        )
        return context

    def get_success_url(self) -> str:
        return super().get_success_url(self.object)


class NozzleDeleteView(NozzleAccessMixin, View):
    def post(self, request, *args, **kwargs) -> HttpResponseRedirect:
        nozzle = self.get_object()
        success_url = self.get_success_url(nozzle)
        nozzle.delete()
        return HttpResponseRedirect(success_url)

class ServiceSessionCreateView(OwnerCompanyMixin, CreateView):
    model = ServiceSession
    form_class = ServiceSessionForm
    template_name = "pages/service_sessions/service_session_start.html"
    # Allow HEAD_ATTENDANT (bombero encargado) to start services
    allowed_roles = ["ADMINISTRATOR", "HEAD_ATTENDANT"]

    def dispatch(self, request, *args, **kwargs):
        """Redirect to the running service if the current branch already has one."""

        profile = getattr(request.user, "profile", None)
        branch_id = getattr(profile, "current_branch_id", None)


        # If the admin user does not have a current branch selected,
        # fall back to the first branch they administrate.
        if branch_id is None and profile:
            admin_branch_ids = get_admin_branch_ids(profile)
            if admin_branch_ids:
                branch_id = admin_branch_ids[0]

        if branch_id:
            active_session = (
                ServiceSession.objects.filter(
                    shift__sucursal_id=branch_id, ended_at__isnull=True
                )
                .order_by("-started_at")
                .first()
            )

            if active_session:
                messages.info(
                    request,
                    "Ya existe un servicio en curso para esta sucursal. "
                    "Redirigiendo al servicio en ejecución.",
                )
                return redirect("service_session_detail", active_session.pk)

        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse("service_session_detail", args=[self.object.pk])


    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        branch_ids = self.get_managed_branch_ids()
        available_shifts = (
            Shift.objects.filter(sucursal_id__in=branch_ids)
            .select_related("sucursal")
            .order_by("sucursal__name", "code")
        )
        # If the current user is a head attendant, limit available shifts
        # to those where they are the configured manager.
        try:
            viewer_profile = getattr(self.request.user, "profile", None)
            if viewer_profile and getattr(viewer_profile, "is_head_ATTENDANT", None) and viewer_profile.is_head_ATTENDANT():
                available_shifts = available_shifts.filter(manager_id=viewer_profile.id)
        except Exception:
            # Defensive: if any attribute access fails, fall back to unfiltered shifts
            pass

        detailed_shifts = available_shifts.select_related(
            "manager__user_FK",
            "manager__position_FK",
        ).prefetch_related("attendants__user_FK", "attendants__position_FK")

        shift_id = (
            self.request.POST.get("shift")
            if self.request.method == "POST"
            else self.request.GET.get("shift")
        )

        selected_shift = None
        if shift_id:
            selected_shift = detailed_shifts.filter(pk=shift_id).first()
        elif self.request.method == "GET":
            selected_shift = detailed_shifts.first()

        self.available_shifts = available_shifts
        self.selected_shift = selected_shift

        kwargs.update(
            {
                "shift": selected_shift,
                "available_shifts": available_shifts,
                "branch_ids": branch_ids,
            }
        )
        return kwargs

    def form_valid(self, form):
        # Ensure head attendants can only start the shift they manage.
        try:
            viewer_profile = getattr(self.request.user, "profile", None)
            if viewer_profile and getattr(viewer_profile, "is_head_ATTENDANT", None) and viewer_profile.is_head_ATTENDANT():
                selected_shift = form.cleaned_data.get("shift")
                if selected_shift and selected_shift.manager_id != viewer_profile.id:
                    form.add_error("shift", "No tienes permiso para iniciar este turno.")
                    messages.error(self.request, "No tienes permiso para iniciar este turno.")
                    return self.form_invalid(form)
        except Exception:
            # Defensive: if anything goes wrong, deny to be safe
            form.add_error("shift", "No es posible verificar permisos para iniciar el turno.")
            messages.error(self.request, "No es posible verificar permisos para iniciar el turno.")
            return self.form_invalid(form)

        response = super().form_valid(form)
        messages.success(
            self.request,
            "Servicio iniciado correctamente. Gestiona la caja e inventario del turno desde esta pantalla.",
        )
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = context.get("form")
        context.update(
            {
                "selected_shift": getattr(self, "selected_shift", None),
                "current_attendants": getattr(form, "current_attendants", []),
                "available_replacements": getattr(form, "available_replacements", []),
                "current_datetime": timezone.localtime(),
                "has_shifts": self.available_shifts.exists(),
            }
        )
        return context


class ServiceSessionDetailView(OwnerCompanyMixin, DetailView):
    model = ServiceSession
    template_name = "pages/service_sessions/service_session_detail.html"
    context_object_name = "service_session"
    # Allow HEAD_ATTENDANT to view/manage the running service details
    allowed_roles = ["ADMINISTRATOR", "HEAD_ATTENDANT", "ATTENDANT"]
    fuel_load_form_prefix = "fuel_load"
    product_load_form_prefix = "product_load"
    product_sale_form_prefix = "product_sale"
    credit_sale_form_prefix = "credit_sale"
    withdrawal_form_prefix = "withdrawal"
    transbank_voucher_form_prefix = "transbank_voucher"
    firefighter_payment_form_prefix = "firefighter_payment"
    close_session_form_prefix = "close_session"

    def dispatch(self, request, *args, **kwargs):
        # Allow an attendant assigned to this specific running service to view it
        try:
            service_pk = kwargs.get("pk")
            if service_pk:
                service = (
                    ServiceSession.objects.filter(pk=service_pk)
                    .prefetch_related("attendants")
                    .first()
                )
                viewer_profile = getattr(request.user, "profile", None)
                if (
                    service
                    and viewer_profile
                    and service.ended_at is None
                    and service.attendants.filter(pk=getattr(viewer_profile, "pk", None)).exists()
                ):
                    # Delegate directly to DetailView to bypass the RoleRequiredMixin
                    return DetailView.dispatch(self, request, *args, **kwargs)
        except Exception:
            # Defensive: fall back to normal dispatch which enforces role checks
            return super().dispatch(request, *args, **kwargs)

        # Default: continue with normal dispatch (enforces role checks / mixins)
        return super().dispatch(request, *args, **kwargs)


    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.ended_at:
            messages.info(
                request,
                "Este servicio ya fue cerrado. Inicia un nuevo servicio para continuar.",
            )
            return redirect("service_session_start")

        context = self.get_context_data(object=self.object)
        return self.render_to_response(context)

    def get_queryset(self):
        branch_ids = self.get_managed_branch_ids()
        queryset = (
            super()
            .get_queryset()
            .select_related(
                "shift__sucursal",
                "shift__manager__user_FK",
                "shift__manager__position_FK",
            )
            .prefetch_related(
                "attendants__user_FK",
                "attendants__position_FK",
                Prefetch(
                    "fuel_loads",
                    queryset=ServiceSessionFuelLoad.objects.select_related(
                        "inventory",
                        "responsible__user_FK",
                    ),
                ),

                Prefetch(
                    "product_loads",
                    queryset=ServiceSessionProductLoad.objects.select_related(
                        "product",
                        "responsible__user_FK",
                    ),
                ),
                Prefetch(
                    "product_sales",
                    queryset=ServiceSessionProductSale.objects.select_related(
                        "responsible__user_FK"
                    ).prefetch_related("items__product"),
                ),
                Prefetch(
                    "credit_sales",
                    queryset=ServiceSessionCreditSale.objects.select_related(
                        "responsible__user_FK",
                        "fuel_inventory",
                    ),
                ),
                Prefetch(
                    "withdrawals",
                    queryset=ServiceSessionWithdrawal.objects.select_related(
                        "responsible__user_FK"
                    ),
                ),
                Prefetch(
                    "transbank_vouchers",
                    queryset=ServiceSessionTransbankVoucher.objects.select_related(
                        "responsible__user_FK"
                    ),
                ),
                Prefetch(
                    "firefighter_payments",
                    queryset=ServiceSessionFirefighterPayment.objects.select_related(
                        "firefighter__user_FK"
                    ),
                ),
            )
        )
        # If the current viewer is an attendant assigned to this specific
        # service (and the service is active), ensure the branch for that
        # service is included so the attendant can access the detail view.
        try:
            viewer_profile = getattr(self.request.user, "profile", None)
            service_pk = self.kwargs.get("pk")
            if (
                viewer_profile
                and service_pk
                and not branch_ids
                and ServiceSession.objects.filter(
                    pk=service_pk, attendants=viewer_profile, ended_at__isnull=True
                ).exists()
            ):
                branch_id = (
                    ServiceSession.objects.filter(pk=service_pk)
                    .values_list("shift__sucursal_id", flat=True)
                    .first()
                )
                if branch_id:
                    branch_ids = list(dict.fromkeys(list(branch_ids) + [branch_id]))

        except Exception:
            pass

        # Allow administrators to access the detail view even if the branch is
        # not part of their managed branch list (e.g., when they just started a
        # service but their profile is not explicitly attached to the branch),
        # but only for the branch they currently have selected.
        try:
            viewer_profile = getattr(self.request.user, "profile", None)
            service_pk = self.kwargs.get("pk")
            if viewer_profile and getattr(viewer_profile, "is_admin", None) and viewer_profile.is_admin():
                if service_pk and not branch_ids:
                    current_branch_id = getattr(viewer_profile, "current_branch_id", None)
                    if current_branch_id and ServiceSession.objects.filter(
                        pk=service_pk, shift__sucursal_id=current_branch_id
                    ).exists():
                        branch_ids = [current_branch_id]
        except Exception:
            pass

        if branch_ids:
            queryset = queryset.filter(shift__sucursal_id__in=branch_ids)
        else:
            queryset = queryset.none()

        return queryset

    def _get_machine_inventory_pairs(self, branch: Sucursal):
        machines = list(
            Machine.objects.filter(island__sucursal=branch)
            .select_related("island", "fuel_inventory")
            .prefetch_related(
                "fuel_inventories",
                "fuel_numerals__fuel_inventory",
                "fuel_numerals__nozzles",
            )
        )

        machine_inventory_pairs = []
        for machine in machines:
            for fuel_inventory in machine.get_fuel_inventories():
                numeral_entries = machine.get_numerals_for_inventory(fuel_inventory)
                for numeral_entry in numeral_entries:
                    machine_inventory_pairs.append(
                        (machine, fuel_inventory, numeral_entry)
                    )
        return machines, machine_inventory_pairs

    def _get_dispense_totals_by_numeral(self) -> dict[int, Decimal]:
        decimal_zero = Decimal("0")
        totals = (
            DispenseEvent.objects.filter(
                service_session=self.object, fuel_numeral_id__isnull=False
            )
            .values("fuel_numeral_id")
            .annotate(
                total_liters=Coalesce(
                    Sum(
                        "litros",
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    ),
                    Value(0, output_field=DecimalField(max_digits=12, decimal_places=2)),
                )
            )
        )
        return {
            entry["fuel_numeral_id"]: Decimal(str(entry.get("total_liters") or decimal_zero))
            for entry in totals
        }

    @staticmethod
    def _group_machine_inventory_forms(machine_inventory_pairs, formset):
        grouped_pairs = []
        combined = list(zip(machine_inventory_pairs, formset.forms))

        current_machine = None
        current_items = []

        for (machine, fuel_inventory, numeral_entry), form in combined:
            if machine != current_machine:
                if current_items:
                    grouped_pairs.append((current_machine, current_items))
                current_machine = machine
                current_items = []
            current_items.append((fuel_inventory, numeral_entry.numeral, form))

        if current_items:
            grouped_pairs.append((current_machine, current_items))

        return grouped_pairs

    @staticmethod
    def _group_machine_nozzle_forms(
        machine_inventory_pairs, formset, dispense_totals_by_numeral=None
    ):
        grouped_pairs = []
        combined = list(zip(machine_inventory_pairs, formset.forms))

        nozzle_lookup = defaultdict(list)
        nozzle_lookup_ids = defaultdict(set)

        dispense_totals_by_numeral = dispense_totals_by_numeral or {}
        decimal_zero = Decimal("0")

        for machine, fuel_inventory, numeral_entry in machine_inventory_pairs:
            for nozzle in machine.nozzles.all():
                fuel_numeral = getattr(nozzle, "fuel_numeral", None)
                if not fuel_numeral:
                    continue
                fuel_numeral_id = fuel_numeral.pk
                key = (fuel_numeral.machine_id, fuel_numeral.fuel_inventory_id, fuel_numeral_id)

                if nozzle.pk in nozzle_lookup_ids[key]:
                    continue
                nozzle_lookup_ids[key].add(nozzle.pk)
                nozzle_lookup[key].append(nozzle)

        current_machine = None
        current_items = []

        for (machine, fuel_inventory, numeral_entry), form in combined:
            if machine != current_machine:
                if current_items:
                    grouped_pairs.append((current_machine, current_items))
                current_machine = machine
                current_items = []
            key = (machine.pk, fuel_inventory.pk, numeral_entry.pk)
            pistol_total = dispense_totals_by_numeral.get(
                numeral_entry.pk, decimal_zero
            )
            current_items.append(
                (
                    fuel_inventory,
                    numeral_entry.numeral,
                    form,
                    nozzle_lookup.get(key, []),
                    pistol_total,
                )
            )

        if current_items:
            grouped_pairs.append((current_machine, current_items))

        return grouped_pairs
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_profile = getattr(self.request.user, "profile", None)
        current_datetime = timezone.localtime()
        fuel_load_form = kwargs.get("fuel_load_form")
        if fuel_load_form is None:
            fuel_load_form = ServiceSessionFuelLoadForm(
                service_session=self.object,
                prefix=self.fuel_load_form_prefix,
            )




        product_load_form = kwargs.get("product_load_form")
        if product_load_form is None:
            product_load_form = ServiceSessionProductLoadForm(
                service_session=self.object,
                prefix=self.product_load_form_prefix,
            )

        product_sale_form = kwargs.get("product_sale_form")
        if product_sale_form is None:
            product_sale_form = ServiceSessionProductSaleForm(
                service_session=self.object,
                responsible_profile=current_profile,
                prefix=self.product_sale_form_prefix,
            )

        credit_sale_form = kwargs.get("credit_sale_form")
        if credit_sale_form is None:
            credit_sale_form = ServiceSessionCreditSaleForm(
                service_session=self.object,
                responsible_profile=current_profile,
                prefix=self.credit_sale_form_prefix,
            )

        product_sale_formset = kwargs.get("product_sale_formset")
        if product_sale_formset is None:
            product_sale_formset = ServiceSessionProductSaleItemFormSet(
                service_session=self.object,
                queryset=ServiceSessionProductSaleItem.objects.none(),
            )

        withdraw_form = kwargs.get("withdraw_form")
        if withdraw_form is None:
            withdraw_form = ServiceSessionWithdrawalForm(
                service_session=self.object,
                responsible_profile=current_profile,
                prefix=self.withdrawal_form_prefix,
            )

        transbank_voucher_form = kwargs.get("transbank_voucher_form")
        if transbank_voucher_form is None:
            transbank_voucher_form = ServiceSessionTransbankVoucherForm(
                service_session=self.object,
                responsible_profile=current_profile,
                prefix=self.transbank_voucher_form_prefix,
            )


        branch = self.object.shift.sucursal
        branch_machines = kwargs.get("branch_machines")
        machine_inventory_pairs = kwargs.get("machine_inventory_pairs")
        if branch_machines is None or machine_inventory_pairs is None:
            branch_machines, machine_inventory_pairs = self._get_machine_inventory_pairs(
                branch
            )
        attendants = list(self.object.attendants.all())
        fuel_loads = list(self.object.fuel_loads.all())
        product_loads = list(self.object.product_loads.all())
        product_sales = list(self.object.product_sales.all())
        for sale in product_sales:
            sale.total_value = sum(
                (
                    item.quantity * item.product.value
                    for item in sale.items.all()
                ),
                Decimal("0"),
            )
        credit_sales = list(self.object.credit_sales.all())
        product_additions = {
            entry["product_id"]: entry["total_added"]
            for entry in self.object.product_loads.values("product_id").annotate(
                total_added=Coalesce(Sum("quantity_added"), 0)
            )
        }
        product_sales_items = {
            entry["product_id"]: entry["total_sold"]
            for entry in ServiceSessionProductSaleItem.objects.filter(
                sale__service_session=self.object
            )
            .values("product_id")
            .annotate(total_sold=Coalesce(Sum("quantity"), 0))
        }
        iot_dispense_events = list(
            DispenseEvent.objects.filter(service_session=self.object)
            .select_related(
                "nozzle__machine__island",
                "nozzle__fuel_numeral__fuel_inventory",
                "fuel_numeral__fuel_inventory",
                "firefighter__user_FK",
            )
            .order_by("-created_at")
        )

        missing_uids = {
            event.uid for event in iot_dispense_events if not event.firefighter and event.uid
        }
        if missing_uids:
            firefighters_by_uid = {
                profile.codigo_identificador: profile
                for profile in Profile.objects.filter(
                    codigo_identificador__in=missing_uids
                ).select_related("user_FK")
            }
            for event in iot_dispense_events:
                if not event.firefighter:
                    event.resolved_firefighter = firefighters_by_uid.get(event.uid)
                else:
                    event.resolved_firefighter = event.firefighter
        else:
            for event in iot_dispense_events:
                event.resolved_firefighter = event.firefighter
        branch_products = list(branch.products.all())
        for product in branch_products:
            product.session_added_quantity = product_additions.get(product.pk, 0)
            product.session_sold_quantity = product_sales_items.get(product.pk, 0)
        withdrawals = list(self.object.withdrawals.all())
        transbank_vouchers = list(self.object.transbank_vouchers.all())
        part_time_attendants = [
            attendant
            for attendant in attendants
            if attendant and not getattr(attendant, "is_partime", True)
        ]
        firefighter_payments = list(self.object.firefighter_payments.all())

        # Determine if the current viewer is a common attendant (ATTENDANT role)
        viewer_profile = getattr(self.request.user, "profile", None)
        is_common_attendant = bool(
            viewer_profile
            and viewer_profile.is_ATTENDANT()
            and not viewer_profile.is_head_ATTENDANT()
        )


        decimal_zero = Decimal("0")
        initial_budget = self.object.initial_budget or decimal_zero
        credit_sales_total = sum(
            (credit.amount or decimal_zero) for credit in credit_sales
        )
        transbank_vouchers_total = sum(
            (voucher.total_amount or decimal_zero) for voucher in transbank_vouchers
        )
        withdrawals_total = sum(
            (withdrawal.amount or decimal_zero) for withdrawal in withdrawals
        )
        product_sales_total = sum(
            (getattr(sale, "total_value", decimal_zero) or decimal_zero)
            for sale in product_sales
        )
        fuel_payments_total = sum(
            (fuel_load.payment_amount or decimal_zero) for fuel_load in fuel_loads
        )
        firefighter_payments_total = sum(
            (payment.amount or decimal_zero) for payment in firefighter_payments
        )
        product_loads_total = sum(
            (product_load.payment_amount or decimal_zero)
            for product_load in product_loads
        )

        close_session_flow_gap = kwargs.get(
            "close_session_flow_gap", self.object.flow_mismatch_amount
        )
        close_session_flow_mismatch_type = kwargs.get(
            "close_session_flow_mismatch_type", self.object.flow_mismatch_type
        )
        flow_mismatch_labels = dict(ServiceSession.FLOW_MISMATCH_CHOICES)

        turn_profit = (
            credit_sales_total
            + transbank_vouchers_total
            + withdrawals_total
            + product_sales_total
        )
        turn_profit_excluding_product_sales = (
            credit_sales_total
            + transbank_vouchers_total
            + withdrawals_total
        )
        net_turn_profit = (
            turn_profit
            - fuel_payments_total
            - firefighter_payments_total
            - product_loads_total
        )

        firefighter_payment_form = kwargs.get("firefighter_payment_form")
        if firefighter_payment_form is None:
            firefighter_payment_form = ServiceSessionFirefighterPaymentForm(
                service_session=self.object,
                firefighters=part_time_attendants,
                prefix=self.firefighter_payment_form_prefix,
            )
        dispense_totals_by_numeral = self._get_dispense_totals_by_numeral()
        close_session_formset = kwargs.get("service_close_formset")
        if close_session_formset is None:
            close_session_formset = ServiceSessionMachineInventoryClosingFormSet(
                prefix=self.close_session_form_prefix,
                machine_inventory_pairs=machine_inventory_pairs,
                pistol_dispense_totals=dispense_totals_by_numeral,
            )
        machine_inventory_close_groups = self._group_machine_inventory_forms(
            machine_inventory_pairs, close_session_formset
        )
        machine_nozzle_close_groups = self._group_machine_nozzle_forms(
            machine_inventory_pairs,
            close_session_formset,
            dispense_totals_by_numeral,
        )
        close_session_flow_details = kwargs.get("close_session_flow_details")
        close_session_flow_total = kwargs.get("close_session_flow_total")
        close_session_flow_missing_prices = kwargs.get(
            "close_session_flow_missing_prices", set()
        )
        context.update(
            {
                "shift": self.object.shift,
                "attendants": self.object.attendants.all(),
                "branch": branch,
                "fuel_inventories": branch.fuel_inventories.all(),
                "fuel_loads": fuel_loads,
                "branch_products": branch_products,
                "product_loads": product_loads,
                "fuel_load_form": fuel_load_form,
                "product_sales": product_sales,
                "product_load_form": product_load_form,
                "fuel_responsible": self.object.shift.manager,
                "product_load_responsible": self.object.shift.manager,
                "product_sale_form": product_sale_form,
                "product_sale_formset": product_sale_formset,
                "credit_sale_form": credit_sale_form,
                "product_responsible": self.object.shift.manager,
                "product_sale_responsible": current_profile,
                "credit_sale_responsible": current_profile,
                "credit_sales": credit_sales,
                "service_date": self.object.started_at.date(),
                "iot_dispense_events": iot_dispense_events,
                "withdrawals": withdrawals,
                "withdraw_form": withdraw_form,
                "withdraw_responsible": current_profile,
                "transbank_voucher_form": transbank_voucher_form,
                "transbank_voucher_responsible": current_profile,
                "transbank_vouchers": transbank_vouchers,
                "current_datetime": current_datetime,
                "current_profile_name": (
                    (current_profile.user_FK.get_full_name() or current_profile.user_FK.username)
                    if current_profile and getattr(current_profile, "user_FK", None)
                    else "Sin encargado asignado"
                ),
                "part_time_attendants": part_time_attendants,
                "firefighter_payment_form": firefighter_payment_form,
                "firefighter_payment_fields": [
                    (
                        attendant,
                        firefighter_payment_form.get_bound_field(attendant),
                    )
                for attendant in part_time_attendants
                ],
                "firefighter_payments": firefighter_payments,
                "service_close_formset": close_session_formset,
                "machine_inventory_close_groups": machine_inventory_close_groups,
                "machine_nozzle_close_groups": machine_nozzle_close_groups,
                "branch_machines": branch_machines,
                "service_session_closed": self.object.ended_at is not None,
                "turn_profit": turn_profit,
                "turn_profit_excluding_product_sales": turn_profit_excluding_product_sales,
                "net_turn_profit": net_turn_profit,
                "close_session_flow_details": close_session_flow_details,
                "close_session_flow_total": close_session_flow_total,
                "close_session_flow_gap": close_session_flow_gap,
                "close_session_flow_mismatch_type": close_session_flow_mismatch_type,
                "close_session_flow_mismatch_label": flow_mismatch_labels.get(
                    close_session_flow_mismatch_type,
                    flow_mismatch_labels[ServiceSession.FLOW_MISMATCH_NONE],
                ),
                "close_session_flow_missing_prices": close_session_flow_missing_prices,
                "turn_profit_components": {
                    "initial_budget": initial_budget,
                    "credit_sales": credit_sales_total,
                    "transbank_vouchers": transbank_vouchers_total,
                    "withdrawals": withdrawals_total,
                    "product_sales": product_sales_total,
                },
                "turn_expenses": {
                    "fuel_loads": fuel_payments_total,
                    "firefighter_payments": firefighter_payments_total,
                    "product_loads": product_loads_total,
                },
                "is_common_attendant": is_common_attendant,
            }
        )
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_type = request.POST.get("form_type", "fuel-load")

        # obtenemos el perfil del usuario logueado
        viewer_profile = getattr(request.user, "profile", None)

        if self.object.ended_at and form_type != "close-session":
            messages.error(
                request, "Este servicio ya fue cerrado y no admite nuevos registros."
            )
            return redirect("service_session_start")
            # 🔒 BLOQUEAR CIERRE DE CAJA A BOMBERO COMÚN
        if form_type == "close-session":
            if viewer_profile and viewer_profile.is_ATTENDANT() and not viewer_profile.is_head_ATTENDANT():
                messages.error(
                    request,
                    "No tienes permisos para realizar el cierre de caja. "
                    "Solo el dueño, administrador o bombero encargado pueden cerrar el turno.",
                )
                return redirect("service_session_detail", pk=self.object.pk)


        if form_type == "close-session":
            close_action = request.POST.get("close_action", "close")
            branch = self.object.shift.sucursal
            if self.object.ended_at:
                messages.info(
                    request,
                    "Este servicio ya fue cerrado previamente.",
                )
                return redirect("service_session_start")
            branch_machines, machine_inventory_pairs = self._get_machine_inventory_pairs(
                branch
            )
            dispense_totals_by_numeral = self._get_dispense_totals_by_numeral()
            close_session_formset = ServiceSessionMachineInventoryClosingFormSet(
                data=request.POST,
                prefix=self.close_session_form_prefix,
                machine_inventory_pairs=machine_inventory_pairs,
                pistol_dispense_totals=dispense_totals_by_numeral,
            )
            if close_session_formset.is_valid():
                machine_inventory_lookup = {

                    (machine.pk, fuel_inventory.pk, numeral_entry.slot): (
                        machine,
                        fuel_inventory,
                        numeral_entry.numeral,
                    )
                    for machine, fuel_inventory, numeral_entry in machine_inventory_pairs
                }
                decimal_zero = Decimal("0")

                if close_action == "check":
                    fuel_prices = {}
                    for price in FuelPrice.objects.filter(sucursal=branch).order_by(
                        "fuel_type", "-created_at", "-pk"
                    ):
                        fuel_prices.setdefault(price.fuel_type, price.price)

                    close_session_flow_total = decimal_zero
                    close_session_flow_details = []
                    missing_price_types = set()

                    for form in close_session_formset:
                        cleaned_data = getattr(form, "cleaned_data", {}) or {}
                        machine_id = cleaned_data.get("machine_id")
                        fuel_inventory_id = cleaned_data.get("fuel_inventory_id")
                        slot = cleaned_data.get("slot")
                        numeral = cleaned_data.get("numeral")
                        if (
                            machine_id is None
                            or fuel_inventory_id is None
                            or slot is None
                            or numeral is None
                        ):
                            continue

                        machine_inventory = machine_inventory_lookup.get(
                            (machine_id, fuel_inventory_id, slot)
                        )
                        if machine_inventory is None:
                            continue

                        machine, fuel_inventory, current_numeral = machine_inventory
                        liters_sold = numeral - current_numeral
                        fuel_type = fuel_inventory.fuel_type if fuel_inventory else None
                        price = fuel_prices.get(fuel_type)

                        if price is None:
                            missing_price_types.add(fuel_type or "Desconocido")
                            flow_amount = decimal_zero
                        else:
                            flow_amount = liters_sold * price

                        close_session_flow_total += flow_amount
                        close_session_flow_details.append(
                            {
                                "machine": machine,
                                "fuel_inventory": fuel_inventory,
                                "liters_sold": liters_sold,
                                "fuel_type": fuel_type,
                                "price": price,
                                "flow_amount": flow_amount,
                            }
                        )

                    if missing_price_types:
                        messages.warning(
                            request,
                            "Falta registrar precio para: "
                            + ", ".join(sorted(missing_price_types)),
                        )

                    context = self.get_context_data(
                        service_close_formset=close_session_formset,
                        branch_machines=branch_machines,
                        machine_inventory_pairs=machine_inventory_pairs,
                        close_session_flow_total=close_session_flow_total,
                        close_session_flow_details=close_session_flow_details,
                        close_session_flow_missing_prices=sorted(missing_price_types),
                    )
                    context["closeSessionModalOpen"] = True


                    turn_profit_excluding_product_sales = (
                        context.get("turn_profit_excluding_product_sales")
                        or decimal_zero
                    )
                    close_session_flow_gap = (
                        turn_profit_excluding_product_sales - close_session_flow_total
                    )
                    close_session_flow_gap = close_session_flow_gap.quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )

                    max_flow_mismatch = Decimal("9999999999.99")
                    if close_session_flow_gap.copy_abs() > max_flow_mismatch:
                        messages.error(
                            request,
                            (
                                "El descuadre calculado excede el límite permitido para "
                                "almacenarlo. Verifica los numerales y montos ingresados."
                            ),
                        )

                        flow_mismatch_labels = dict(
                            ServiceSession.FLOW_MISMATCH_CHOICES
                        )
                        context.update(
                            {
                                "close_session_flow_gap": close_session_flow_gap,
                                "close_session_flow_mismatch_type": ServiceSession.FLOW_MISMATCH_NONE,
                                "close_session_flow_mismatch_label": flow_mismatch_labels[
                                    ServiceSession.FLOW_MISMATCH_NONE
                                ],
                            }
                        )
                        return self.render_to_response(context)


                    if close_session_flow_gap > decimal_zero:
                        flow_mismatch_type = ServiceSession.FLOW_MISMATCH_POSITIVE
                    elif close_session_flow_gap < decimal_zero:
                        flow_mismatch_type = ServiceSession.FLOW_MISMATCH_NEGATIVE
                    else:
                        flow_mismatch_type = ServiceSession.FLOW_MISMATCH_NONE

                    self.object.flow_mismatch_amount = close_session_flow_gap
                    self.object.flow_mismatch_type = flow_mismatch_type
                    self.object.save(
                        update_fields=["flow_mismatch_amount", "flow_mismatch_type"]
                    )

                    flow_mismatch_labels = dict(ServiceSession.FLOW_MISMATCH_CHOICES)
                    context.update(
                        {
                            "close_session_flow_gap": close_session_flow_gap,
                            "close_session_flow_mismatch_type": flow_mismatch_type,
                            "close_session_flow_mismatch_label": flow_mismatch_labels.get(
                                flow_mismatch_type,
                                flow_mismatch_labels[ServiceSession.FLOW_MISMATCH_NONE],
                            ),
                        }
                    )
                    return self.render_to_response(context)

                closure_time = timezone.now()
                with transaction.atomic():
                    for form in close_session_formset:
                        machine_id = form.cleaned_data.get("machine_id")
                        fuel_inventory_id = form.cleaned_data.get("fuel_inventory_id")
                        slot = form.cleaned_data.get("slot")
                        numeral = form.cleaned_data.get("numeral")
                        if (
                            machine_id is None
                            or fuel_inventory_id is None
                            or slot is None

                            or numeral is None
                        ):
                            continue
                        machine_inventory = machine_inventory_lookup.get(
                            (machine_id, fuel_inventory_id, slot)
                        )
                        if machine_inventory is None:
                            continue

                        machine, fuel_inventory, current_numeral = machine_inventory
                        liters_sold = numeral - current_numeral
                        if fuel_inventory_id and liters_sold > decimal_zero:
                            FuelInventory.objects.filter(pk=fuel_inventory_id).update(
                                liters=F("liters") - liters_sold
                            )
                        MachineFuelInventoryNumeral.objects.update_or_create(
                            machine=machine,
                            fuel_inventory=fuel_inventory,
                            slot=slot,
                            defaults={"numeral": numeral},
                        )
                    ServiceSession.objects.filter(
                        shift__sucursal=self.object.shift.sucursal,
                        ended_at__isnull=True,
                    ).exclude(pk=self.object.pk).update(ended_at=closure_time)
                    self.object.ended_at = closure_time
                    self.object.save(update_fields=["ended_at"])
                messages.success(
                    request,
                    "Caja cerrada y servicio finalizado correctamente.",
                )
                return redirect("service_session_start")

            context = self.get_context_data(
                service_close_formset=close_session_formset,
                branch_machines=branch_machines,
                machine_inventory_pairs=machine_inventory_pairs,
            )
            return self.render_to_response(context)

        if form_type == "product-load":
            form = ServiceSessionProductLoadForm(
                data=request.POST,
                service_session=self.object,
                prefix=self.product_load_form_prefix,
            )
            if form.is_valid():
                form.save()
                messages.success(
                    request,
                    "Ingreso de productos registrado correctamente.",
                )
                return redirect("service_session_detail", pk=self.object.pk)

            context = self.get_context_data(product_load_form=form)
            return self.render_to_response(context)
        if form_type == "product-sale":
            sale_form = ServiceSessionProductSaleForm(
                data=request.POST,
                service_session=self.object,
                responsible_profile=getattr(request.user, "profile", None),
                prefix=self.product_sale_form_prefix,
            )
            item_formset = ServiceSessionProductSaleItemFormSet(
                data=request.POST,
                service_session=self.object,
                queryset=ServiceSessionProductSaleItem.objects.none(),
            )
            if sale_form.is_valid() and item_formset.is_valid():
                with transaction.atomic():
                    sale = sale_form.save()
                    for form in item_formset:
                        if not form.cleaned_data or form.cleaned_data.get("DELETE"):
                            continue
                        product = form.cleaned_data["product"]
                        quantity = form.cleaned_data["quantity"]
                        ServiceSessionProductSaleItem.objects.create(
                            sale=sale,
                            product=product,
                            quantity=quantity,
                        )
                        BranchProduct.objects.filter(pk=product.pk).update(
                            quantity=F("quantity") - quantity
                        )
                    messages.success(
                        request,
                        "Venta de productos registrada correctamente.",
                    )
                return redirect("service_session_detail", pk=self.object.pk)

            context = self.get_context_data(
                product_sale_form=sale_form,
                product_sale_formset=item_formset,
            )
            return self.render_to_response(context)

        if form_type == "credit-sale":
            credit_form = ServiceSessionCreditSaleForm(
                data=request.POST,
                service_session=self.object,
                responsible_profile=getattr(request.user, "profile", None),
                prefix=self.credit_sale_form_prefix,
            )
            if credit_form.is_valid():
                credit_form.save()
                messages.success(
                    request,
                    "Venta a crédito registrada correctamente.",
                )
                return redirect("service_session_detail", pk=self.object.pk)

            context = self.get_context_data(credit_sale_form=credit_form)
            return self.render_to_response(context)

        if form_type == "withdrawal":
            withdraw_form = ServiceSessionWithdrawalForm(
                data=request.POST,
                service_session=self.object,
                responsible_profile=getattr(request.user, "profile", None),
                prefix=self.withdrawal_form_prefix,
            )
            if withdraw_form.is_valid():
                withdraw_form.save()
                messages.success(
                    request,
                    "Tirada de caja registrada correctamente.",
                )
                return redirect("service_session_detail", pk=self.object.pk)

            context = self.get_context_data(withdraw_form=withdraw_form)
            return self.render_to_response(context)

        if form_type == "transbank-voucher":
            voucher_form = ServiceSessionTransbankVoucherForm(
                data=request.POST,
                service_session=self.object,
                responsible_profile=getattr(request.user, "profile", None),
                prefix=self.transbank_voucher_form_prefix,
            )
            if voucher_form.is_valid():
                voucher_form.save()
                messages.success(
                    request,
                    "Registro de vouchers de Transbank guardado correctamente.",
                )
                return redirect("service_session_detail", pk=self.object.pk)

            context = self.get_context_data(transbank_voucher_form=voucher_form)
            return self.render_to_response(context)

        if form_type == "firefighter-payment":
            part_time_attendants = [
                attendant
                for attendant in self.object.attendants.all()
                if attendant and not getattr(attendant, "is_partime", True)
            ]
            firefighter_payment_form = ServiceSessionFirefighterPaymentForm(
                data=request.POST,
                service_session=self.object,
                firefighters=part_time_attendants,
                prefix=self.firefighter_payment_form_prefix,
            )
            if firefighter_payment_form.is_valid():
                firefighter_payment_form.save()
                messages.success(
                    request,
                    "Pagos a bomberos registrados correctamente.",
                )
                return redirect("service_session_detail", pk=self.object.pk)

            context = self.get_context_data(
                firefighter_payment_form=firefighter_payment_form
            )
            return self.render_to_response(context)
        form = ServiceSessionFuelLoadForm(
            data=request.POST,
            service_session=self.object,
            prefix=self.fuel_load_form_prefix,
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Carga de combustible registrada correctamente.")
            return redirect("service_session_detail", pk=self.object.pk)

        context = self.get_context_data(fuel_load_form=form)
        return self.render_to_response(context)